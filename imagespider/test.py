# -*- coding: utf-8 -*-
from collections import OrderedDict
from openpyxl import Workbook
import json,re,time,urlparse,codecs

# log文件

file = codecs.open("/Users/apple/Desktop/access.log-20170624",'r',encoding = "ISO-8859-1")
# file = codecs.open("/Users/apple/Desktop/error1.txt",'r',encoding = "ISO-8859-1")

# JSON文件
jsonFile = open('/Users/apple/Desktop/log.json', 'a')
# error文件
error = open('/Users/apple/Desktop/error.txt', 'a')
# excl
excl = '/Users/apple/Desktop/log.xlsx'
# 计数
i = 0

#excel
wb = Workbook()
ws = wb.active
ws.title = "log"

def getRefer(url):
    result = urlparse.urlparse(url)
    params = urlparse.parse_qs(result.query, False)
    if params.has_key('refer'):
        return params['refer'][0]
    else:
        return ''


def isMobile(user_agent):
    ret = re.findall(
        r'(android|bb\d+|meego).+mobile|avantgo|bada\/|blackberry|blazer|compal|elaine|fennec|hiptop|iemobile|ip(hone|od)|iris|kindle|lge |maemo|midp|mmp|mobile.+firefox|netfront|opera m(ob|in)i|palm( os)?|phone|p(ixi|re)\/|plucker|pocket|psp|series(4|6)0|symbian|treo|up\.(browser|link)|vodafone|wap|windows ce|xda|xiino',
        user_agent.lower())
    return ret==True


def formatTiem(localtime):
    timeArray = time.strptime(localtime, '%d/%b/%Y:%H:%M:%S +0800')
    otherStyleTime = time.strftime("%Y/%m/%d %H:%M:%S", timeArray)
    return otherStyleTime

def wirteExcel(log):
    ws['A%d'%i] = log['ip1']
    ws['B%d'%i] = log['ip2']
    ws['C%d'%i] = log['date']
    ws['D%d'%i] = log['method']
    ws['E%d'%i] = log['url']
    ws['F%d'%i] = log['referer']
    ws['G%d'%i] = log['user_agent']
    ws['H%d'%i] = log['url_refer']
    ws['I%d'%i] = log['referer_refer']
    ws['J%d'%i] = log['is_mobile']

def log2Json(line):
    log = OrderedDict()
    data = line.split(' ')
    log['ip1'] = data[0]
    log['ip2'] = data[-0].replace('"', '')
    log['date'] = formatTiem((data[3] + ' ' + data[4]).replace('[', '').replace(']', ''))
    log['method'] = data[5].replace('"', '')
    log['url'] = data[6]
    log['referer'] = data[10].replace('"', '')
    log['user_agent'] = line.split('" "')[1]

    log['url_refer'] = getRefer(log['url'])
    log['referer_refer'] = getRefer(log['referer'])
    log['is_mobile'] = isMobile(log['user_agent'])

    if log['url'] == '-':
        log['url'] = ''

    if log['referer'] == '-':
        log['referer'] = ''

    if log['ip2'] == '-':
        log['ip2'] = ''

    if log['user_agent'] == '-':
        log['user_agent'] = ''

    wirteExcel(log)

    jsonStr = json.dumps(log)

    if (i != 0):
        jsonStr = ',' + jsonStr
    jsonFile.write(jsonStr)

    print line


if __name__ == '__main__':
    jsonFile.write('[')

    for line in file:
        try:
            log2Json(line)
        except Exception as err:
            error.write(line)
            print(err)
        i+=1

        # if i==10:
        #     break


    jsonFile.write(']')
    jsonFile.close()
    file.close()
    error.close()
    wb.save(excl)
    print 'ok %d' % i
